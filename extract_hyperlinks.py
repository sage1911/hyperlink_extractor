import openpyxl
import re
import pandas as pd

file_path = "workflow31mar.xlsm"  # Replace with your actual file path

# Load workbook and allow user to select a sheet or apply to all
wb = openpyxl.load_workbook(file_path, data_only=False)  # Read formulas

sheet_names = wb.sheetnames
print("Available sheets:", sheet_names)
selected_sheet = input("Enter sheet name to process (or type 'ALL' to process all sheets): ")


def process_sheet(ws):
    # Read header row
    header = [cell.value for cell in ws[1]]  # Extract column names

    try:
        doc_col_idx = header.index("Document Number")  # Get zero-based index
        user_col_idx = header.index("RQC User")  # Get zero-based index
    except ValueError:
        print("Required columns not found in sheet", ws.title)
        return []

    # Optional columns that we want to include if they exist
    optional_columns = ["Study", "Study Country", "Study Site", "Document Name"]
    optional_col_indices = {
        col: header.index(col) for col in optional_columns if col in header
    }  # Store indices of existing columns

    # Extract hyperlinks from "Document Number"
    hyperlink_data = []
    for row in ws.iter_rows(min_row=2, values_only=False):  # Iterate over rows (excluding header)
        doc_cell = row[doc_col_idx]  # Get the "Document Number" cell

        if doc_cell.data_type == "f":  # Check if the cell contains a formula
            formula = doc_cell.value  # Get the formula text
            # Original code assumes formula is a string suitable for regex if type is 'f'
            match = re.search(r'HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', formula)  # Extract hyperlink & identifier
            if match:
                data_row = {"Identifier": match.group(2), "URL": match.group(1)}  # Store Identifier & URL
                for col_name, col_idx in optional_col_indices.items():  # Extract optional columns if present
                    try:
                        data_row[col_name] = row[col_idx].value if row[col_idx] else None # Get column value
                    except IndexError:
                        data_row[col_name] = None # If column is missing, set to None
                hyperlink_data.append(data_row)
    return hyperlink_data


all_hyperlink_data = []
if selected_sheet.upper() == "ALL":
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        all_hyperlink_data.extend(process_sheet(ws))
else:
    ws = wb[selected_sheet]
    all_hyperlink_data.extend(process_sheet(ws))

# Convert to DataFrame
df_hyperlinks = pd.DataFrame(all_hyperlink_data)

# Load workbook again to get evaluated values
wb = openpyxl.load_workbook(file_path, data_only=True)  # Read evaluated values


def process_rqc(ws):
    rqc_data = []
    header = [cell.value for cell in ws[1]]  # Extract column names

    try:
        doc_col_idx = header.index("Document Number")  # Get zero-based index
        user_col_idx = header.index("RQC User")  # Get zero-based index
    except ValueError:
        print("Required columns not found in sheet", ws.title)
        return []

    for row in ws.iter_rows(min_row=2, values_only=True):  # Iterate over rows (excluding header)
        # Original code assumes row has enough elements for indices
        identifier = row[doc_col_idx]  # Get Document Number displayed text
        rqc_user = row[user_col_idx]  # Get evaluated RQC User value
        rqc_data.append((identifier, rqc_user))
    return rqc_data


all_rqc_data = []
if selected_sheet.upper() == "ALL":
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        all_rqc_data.extend(process_rqc(ws))
else:
    ws = wb[selected_sheet]
    all_rqc_data.extend(process_rqc(ws))

df_rqc = pd.DataFrame(all_rqc_data, columns=["Identifier", "RQC User"])

# Merge data
if not df_hyperlinks.empty:
    df_merged = df_hyperlinks.merge(df_rqc, on="Identifier", how="inner")
    df_filtered = df_merged[df_merged["RQC User"] == "Vasuki"] # Filter for Vasuki

    # Save to Excel
    output_file = "extracted_links.xlsx"
    df_filtered.to_excel(output_file, index=False)
    print(f"Extracted {len(df_filtered)} links saved to {output_file}")
else:
    print("No hyperlinks extracted.")